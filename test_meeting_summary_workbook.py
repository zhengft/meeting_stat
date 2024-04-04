#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from datetime import time
from functools import partial
from itertools import islice
from operator import attrgetter, itemgetter
from pprint import pprint

import pytest
from openpyxl import Workbook

from meeting_comm import Cell, cross, pipe, swap_args
from meeting_attendance_workbook import (
    AttendanceInfo, create_origin_attendance_info
)
from meeting_summary_workbook import (
    GET_NAME_ERROR, GET_GROUP_ERROR, GET_TEAM_ERROR, GET_NUMBER_ERROR,
    MEETING_INFO_SHEET_NAME, PEOPLE_SHEET_NAME,
    MeetingInfo,
    ParsePersonnelInfoError, PersoneelInfo, TeamAttendanceInfo, TeamLocation,
    contains_formal_name,
    filter_infos_by_group_team,
    filter_unmatched_attendance_infos,
    generate_present_commands,
    generate_team_commands,
    generate_title_command,
    get_group_teams_by_personeel_infos,
    get_groups_by_personeel_infos,
    group_attendance_sheet_to_team_locations,
    group_infos_by_group,
    group_infos_by_team,
    is_enough_attendance_time,
    is_person_present,
    item_extract,
    calc_group_attendance_infos,
    parse_meeting_info,
    parse_meeting_info_sheet,
    parse_personnel_info, parse_people_sheet,
    partition_by_time,
    partition_infos,
    partition_present,
    partition_present_personeel_infos,
)
from test_meeting_attendance_workbook import (
    TEST_ATTENDANCE_INFO_01, TEST_ATTENDANCE_INFO_01_01,
    TEST_ATTENDANCE_INFO_02, TEST_ATTENDANCE_INFO_03, TEST_ATTENDANCE_INFO_04,
    get_test_attendance_infos_01
)


def test_parse_personnel_info_01():
    row = (
        Cell(1), Cell('无门'), Cell('元'), Cell('中乾'),
        Cell('中乾0'), Cell('组长')
    )
    result = parse_personnel_info(row)
    expected = PersoneelInfo(name='无门', group='元', team='中乾', number=0)
    assert expected == result


def test_parse_personnel_info_02():
    row = (
        Cell(1), Cell('无门'), Cell('元'), Cell('中乾'),
        Cell('中乾'), Cell('组长')
    )
    with pytest.raises(ParsePersonnelInfoError) as exp:
        parse_personnel_info(row)

    assert GET_NUMBER_ERROR == str(exp.value)


def test_parse_personnel_info_03():
    row = (
        Cell(1), Cell(''), Cell('元'), Cell('中乾'),
        Cell('中乾0'), Cell('组长')
    )
    with pytest.raises(ParsePersonnelInfoError) as exp:
        parse_personnel_info(row)

    assert GET_NAME_ERROR == str(exp.value)


def test_parse_personnel_info_04():
    row = (
        Cell(1), Cell('无门'), Cell(''), Cell('中乾'),
        Cell('中乾0'), Cell('组长')
    )
    with pytest.raises(ParsePersonnelInfoError) as exp:
        parse_personnel_info(row)

    assert GET_GROUP_ERROR == str(exp.value)


def test_parse_personnel_info_05():
    row = (
        Cell(1), Cell('无门'), Cell('元'), Cell(''),
        Cell('中乾0'), Cell('组长')
    )
    with pytest.raises(ParsePersonnelInfoError) as exp:
        parse_personnel_info(row)

    assert GET_TEAM_ERROR == str(exp.value)


TEST_GROUP_NAME_01 = '大组0'
TEST_GROUP_NAME_02 = '大组1'


def create_test_summary_workbook_01() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = PEOPLE_SHEET_NAME
    ws.cell(row=1, column=1, value='序号')
    ws.cell(row=1, column=2, value='姓名')
    ws.cell(row=1, column=3, value='大组')
    ws.cell(row=1, column=4, value='小组')
    ws.cell(row=1, column=5, value='编号')
    ws.cell(row=1, column=6, value='组长')
    ws.cell(row=1, column=7, value='退出')

    ws.cell(row=2, column=1, value=1)
    ws.cell(row=2, column=2, value='人员1')
    ws.cell(row=2, column=3, value='大组0')
    ws.cell(row=2, column=4, value='中乾')
    ws.cell(row=2, column=5, value='小组编号0')
    ws.cell(row=2, column=6, value='组长')

    ws.cell(row=3, column=1, value=2)
    ws.cell(row=3, column=2, value='人员2')
    ws.cell(row=3, column=3, value='大组0')
    ws.cell(row=3, column=4, value='中乾')
    ws.cell(row=3, column=5, value='小组编号1')

    ws.cell(row=4, column=1, value=3)
    ws.cell(row=4, column=2, value='人员3')
    ws.cell(row=4, column=3, value='大组0')
    ws.cell(row=4, column=4, value='中乾')
    ws.cell(row=4, column=5, value='小组编号2')

    ws.cell(row=5, column=1, value=4)
    ws.cell(row=5, column=2, value='人员4')
    ws.cell(row=5, column=3, value='大组0')
    ws.cell(row=5, column=4, value='中坤')
    ws.cell(row=5, column=5, value='小组编号0')
    ws.cell(row=5, column=6, value='组长')

    ws.cell(row=6, column=1, value=5)
    ws.cell(row=6, column=2, value='人员5')
    ws.cell(row=6, column=3, value='大组0')
    ws.cell(row=6, column=4, value='中坤')
    ws.cell(row=6, column=5, value='小组编号1')

    ws.cell(row=7, column=1, value=6)
    ws.cell(row=7, column=2, value='')
    ws.cell(row=7, column=3, value='大组0')
    ws.cell(row=7, column=4, value='中坤')
    ws.cell(row=7, column=5, value='小组编号2')

    ws.cell(row=8, column=1, value=7)
    ws.cell(row=8, column=2, value='人员7')
    ws.cell(row=8, column=3, value='大组1')
    ws.cell(row=8, column=4, value='上乾')
    ws.cell(row=8, column=5, value='小组编号0')
    ws.cell(row=8, column=6, value='组长')

    ws.cell(row=9, column=1, value=8)
    ws.cell(row=9, column=2, value='')
    ws.cell(row=9, column=3, value='大组1')
    ws.cell(row=9, column=4, value='中乾')
    ws.cell(row=9, column=5, value='小组编号1')

    ws.cell(row=10, column=1, value=9)
    ws.cell(row=10, column=2, value=None)
    ws.cell(row=10, column=3, value='大组1')
    ws.cell(row=10, column=4, value='中乾')
    ws.cell(row=10, column=5, value='小组编号2')

    ws = wb.create_sheet(MEETING_INFO_SHEET_NAME)
    ws.cell(row=1, column=1, value='节气名')
    ws.cell(row=1, column=2, value='冬至')
    ws.cell(row=2, column=1, value='会议总时长')
    ws.cell(row=2, column=2, value='75')

    ws = wb.create_sheet(TEST_GROUP_NAME_01)
    ws.cell(row=1, column=1, value='中乾组（3人）')
    ws.cell(row=1, column=2, value='序号')
    ws.cell(row=1, column=3, value='用户入会昵称')
    ws.cell(row=1, column=4, value='累计参会时长')
    ws.cell(row=2, column=1, value='')
    ws.cell(row=2, column=2, value='')
    ws.cell(row=2, column=3, value='')
    ws.cell(row=2, column=4, value='')
    ws.cell(row=3, column=1, value='')
    ws.cell(row=3, column=2, value='')
    ws.cell(row=3, column=3, value='')
    ws.cell(row=3, column=4, value='')
    ws.cell(row=4, column=1, value='')
    ws.cell(row=4, column=2, value='')
    ws.cell(row=4, column=3, value='')
    ws.cell(row=4, column=4, value='')

    ws.cell(row=5, column=1, value='中坤组（3人）')
    ws.cell(row=5, column=2, value='序号')
    ws.cell(row=5, column=3, value='用户入会昵称')
    ws.cell(row=5, column=4, value='累计参会时长')

    return wb


TEST_SUMMARY_WB_01 = create_test_summary_workbook_01()

TEST_PERSONEEL_INFO_01 = PersoneelInfo(
    name='人员1', group='大组0', team='中乾', number=0
)
TEST_PERSONEEL_INFO_02 = PersoneelInfo(
    name='人员2', group='大组0', team='中乾', number=1
)
TEST_PERSONEEL_INFO_03 = PersoneelInfo(
    name='人员3', group='大组0', team='中乾', number=2
)
TEST_PERSONEEL_INFO_04 = PersoneelInfo(
    name='人员4', group='大组0', team='中坤', number=0
)
TEST_PERSONEEL_INFO_05 = PersoneelInfo(
    name='人员5', group='大组0', team='中坤', number=1
)
TEST_PERSONEEL_INFO_07 = PersoneelInfo(
    name='人员7', group='大组1', team='上乾', number=0
)

TEST_MEETING_INFO_01 = parse_meeting_info_sheet(
    TEST_SUMMARY_WB_01[MEETING_INFO_SHEET_NAME]
)

TEST_GROUP_ATTENDANCE_SHEET_01 = TEST_SUMMARY_WB_01[TEST_GROUP_NAME_01]

# TEST_TEAM_ATTENDANCES_01 = {
#     '中乾': (('大组0', '中乾'), (), (), ()),
#     '中坤': (('大组0', '中坤'), (), (), ()),
# }

TEST_TEAM_ATTENDANCE_INFO_01 = TeamAttendanceInfo(
    ('大组0', '中乾'),
    ((TEST_PERSONEEL_INFO_03, TEST_ATTENDANCE_INFO_03),),
    ((TEST_PERSONEEL_INFO_01, TEST_ATTENDANCE_INFO_01),),
    (TEST_PERSONEEL_INFO_02,),
)

TEST_TEAM_ATTENDANCE_INFO_02 = TeamAttendanceInfo(
    ('大组0', '中坤'),
    ((TEST_PERSONEEL_INFO_04, TEST_ATTENDANCE_INFO_04),),
    (),
    (TEST_PERSONEEL_INFO_05,),
)

TEST_TEAM_ATTENDANCE_INFOS_01 = {
    '中乾': TEST_TEAM_ATTENDANCE_INFO_01,
    '中坤': TEST_TEAM_ATTENDANCE_INFO_02,
}

TEST_TEAM_LOCATION_01 = TeamLocation('中乾', 1)


get_test_personeel_infos_01 = pipe(
    partial(
        parse_people_sheet, TEST_SUMMARY_WB_01[PEOPLE_SHEET_NAME]
    ),
    itemgetter(0),
)


def test_contains_formal_name_01():
    result = contains_formal_name((
        TEST_PERSONEEL_INFO_01, TEST_ATTENDANCE_INFO_01
    ))
    assert result is True


def test_contains_formal_name_02():
    result = contains_formal_name((
        TEST_PERSONEEL_INFO_01, TEST_ATTENDANCE_INFO_01_01
    ))
    assert result is True


def test_contains_formal_name_03():
    result = contains_formal_name((
        TEST_PERSONEEL_INFO_01, TEST_ATTENDANCE_INFO_02
    ))
    assert result is False


def test_parse_people_sheet_01():
    result = parse_people_sheet(TEST_SUMMARY_WB_01[PEOPLE_SHEET_NAME])
    expected = (
        (
            TEST_PERSONEEL_INFO_01,
            TEST_PERSONEEL_INFO_02,
            TEST_PERSONEEL_INFO_03,
            TEST_PERSONEEL_INFO_04,
            TEST_PERSONEEL_INFO_05,
            TEST_PERSONEEL_INFO_07,
        ),
        ((0, '中乾'), (1, '中坤'), (2, '上乾'), (3, '中乾')),
    )
    assert expected == result


def test_is_person_present_01():
    result = is_person_present((
        PersoneelInfo(name='人员1', group='大组0', team='中乾', number=0),
        get_test_attendance_infos_01()
    ))
    assert result is True


def test_is_person_present_02():
    result = is_person_present((
        PersoneelInfo(name='人员7', group='大组1', team='上乾', number=0),
        get_test_attendance_infos_01()
    ))
    assert result is False


def test_partition_present_personeel_infos_01():
    result = partition_present_personeel_infos(
        (get_test_personeel_infos_01(), get_test_attendance_infos_01())
    )
    expected_lens = (3, 3)
    expected_names = (
        ('人员1', '人员3', '人员4'), ('人员2', '人员5', '人员7')
    )
    assert expected_lens == tuple(map(len, result))
    assert expected_names == tuple(
        map(pipe(partial(map, attrgetter('name')), tuple), result)
    )


def test_filter_unmatched_attendance_infos_01():
    result = filter_unmatched_attendance_infos(
        get_test_personeel_infos_01(), get_test_attendance_infos_01()
    )
    assert 1 == len(result)


def test_parse_meeting_info_01():
    result = parse_meeting_info((Cell('节气名'), Cell('冬至')))
    expected = ('solar_term', '冬至')
    assert expected == result


def test_parse_meeting_info_02():
    result = parse_meeting_info((Cell('会议总时长'), Cell('75')))
    expected = ('meeting_time', time(0, 40))
    assert expected == result


def test_parse_meeting_info_sheet_01():
    result = parse_meeting_info_sheet(TEST_SUMMARY_WB_01[MEETING_INFO_SHEET_NAME])
    expected = MeetingInfo('冬至', time(0, 40))
    assert expected == result


def test_is_enough_attendance_time_01():
    result = is_enough_attendance_time(
        TEST_MEETING_INFO_01,
        TEST_ATTENDANCE_INFO_01,
    )
    assert result is False


def test_is_enough_attendance_time_02():
    result = is_enough_attendance_time(
        TEST_MEETING_INFO_01,
        create_origin_attendance_info(
            '人员1', '人员1', time(0, 40, 0),
        ),
    )
    assert result is True


def test_is_enough_attendance_time_03():
    result = is_enough_attendance_time(
        TEST_MEETING_INFO_01,
        create_origin_attendance_info(
            '人员1', '人员1', time(0, 39, 59),
        ),
    )
    assert result is False


def test_partition_present_01():
    present, absent_people = partition_present(
        get_test_personeel_infos_01(),
        get_test_attendance_infos_01()
    )
    assert 3 == len(present)
    assert 3 == len(absent_people)


def test_partition_infos_01():
    a, b, c = partition_infos(
        TEST_MEETING_INFO_01,
        get_test_personeel_infos_01(),
        get_test_attendance_infos_01()
    )
    assert 2 == len(a)
    assert 1 == len(b)
    assert 3 == len(c)


def test_get_group_teams_by_personeel_infos_01():
    result = get_group_teams_by_personeel_infos(get_test_personeel_infos_01())
    expected = (
        ('大组0', '中乾'),
        ('大组0', '中坤'),
        ('大组1', '上乾'),
    )
    assert expected == result


def test_get_groups_by_personeel_infos_01():
    result = get_groups_by_personeel_infos(get_test_personeel_infos_01())
    expected = ('大组0', '大组1')
    assert expected == result


def test_filter_infos_by_group_team_01():
    result = filter_infos_by_group_team(
        ('元', '中乾'),
        *islice(
            partition_infos(
                TEST_MEETING_INFO_01,
                get_test_personeel_infos_01(),
                get_test_attendance_infos_01()
            ),
            3,
        )
    )
    assert ('元', '中乾') == result[0]
    assert 1 == len(result[1])
    assert 1 == len(result[2])
    assert 1 == len(result[3])
    # pprint(result)


def test_group_infos_by_group_01():
    result = group_infos_by_group(
        (('大组0', '中乾'), (), (), ()),
        (('大组0', '中坤'), (), (), ()),
        (('大组1', '上乾'), (), (), ()),
    )
    expected = {
        '大组0': (
            (('大组0', '中乾'), (), (), ()),
            (('大组0', '中坤'), (), (), ()),
        ),
        '大组1': (
            (('大组1', '上乾'), (), (), ()),
        ),
    }
    assert expected == result


def test_group_infos_by_team_01():
    result = group_infos_by_team(
        (('大组0', '中乾'), (), (), ()),
        (('大组0', '中坤'), (), (), ()),
        (('大组1', '上乾'), (), (), ()),
    )
    expected = {
        '中乾': (('大组0', '中乾'), (), (), ()),
        '中坤': (('大组0', '中坤'), (), (), ()),
        '上乾': (('大组1', '上乾'), (), (), ()),
    }
    assert expected == result


def test_calc_group_attendance_infos_01():
    result_groups = calc_group_attendance_infos(
        TEST_MEETING_INFO_01,
        get_test_personeel_infos_01(),
        get_test_attendance_infos_01()
    )
    assert 2 == len(result_groups)
    assert ('大组0', '大组1') == tuple(result_groups)
    assert ('中乾', '中坤') == tuple(result_groups['大组0'])
    assert ('大组0', '中乾') == result_groups['大组0']['中乾'][0]
    assert 1 == len(result_groups['大组0']['中乾'][1])
    assert 1 == len(result_groups['大组0']['中乾'][2])
    assert 1 == len(result_groups['大组0']['中乾'][3])
    assert ('大组0', '中坤') == result_groups['大组0']['中坤'][0]
    assert 1 == len(result_groups['大组0']['中坤'][1])
    assert 0 == len(result_groups['大组0']['中坤'][2])
    assert 1 == len(result_groups['大组0']['中坤'][3])
    assert ('上乾',) == tuple(result_groups['大组1'])
    assert ('大组1', '上乾') == result_groups['大组1']['上乾'][0]
    assert 0 == len(result_groups['大组1']['上乾'][1])
    assert 0 == len(result_groups['大组1']['上乾'][2])
    assert 1 == len(result_groups['大组1']['上乾'][3])
    # pprint(result)

def test_group_attendance_sheet_to_team_locations_01():
    result = group_attendance_sheet_to_team_locations(
        TEST_GROUP_ATTENDANCE_SHEET_01
    )
    expected = (('中乾', 1), ('中坤', 5))
    assert expected == result


def test_generate_present_commands_01():
    result = generate_present_commands(
        TEST_TEAM_LOCATION_01, (
            (TEST_PERSONEEL_INFO_01, TEST_ATTENDANCE_INFO_01),
            (TEST_PERSONEEL_INFO_03, TEST_ATTENDANCE_INFO_03),
        ),
        (True, False),
    )
    expected = (
        (2, 2, 1, False),
        (2, 3, '中乾0人员1', False),
        (2, 4, '00:30:00', True),
        (3, 2, 2, False),
        (3, 3, '中乾2人员3＆中坤0人员4', False),
        (3, 4, '00:40:00', False),
    )
    assert expected == result


def test_generate_title_command_01():
    result = generate_title_command(
        TEST_TEAM_LOCATION_01, TEST_TEAM_ATTENDANCE_INFO_01
    )
    expected = (1, 1, '中乾组（3人）', False)
    assert expected == result


def test_generate_team_commands_01():
    result = generate_team_commands(
        TEST_TEAM_LOCATION_01, TEST_TEAM_ATTENDANCE_INFOS_01
    )
    expected = (
        (1, 1, '中乾组（3人）', False),
        (2, 2, 1, False),
        (2, 3, '中乾0人员1', False),
        (2, 4, '00:30:00', True),
        (3, 2, 2, False),
        (3, 3, '中乾2人员3＆中坤0人员4', False),
        (3, 4, '00:40:00', False),
        (2, 1, '中乾0人员1', False),
        (3, 1, '中乾1人员2', False),
    )
    assert expected == result


def test_item_extract_01():
    result = item_extract(
        '中乾', TEST_TEAM_ATTENDANCE_INFOS_01
    )
    expected = TEST_TEAM_ATTENDANCE_INFO_01
    assert expected == result

