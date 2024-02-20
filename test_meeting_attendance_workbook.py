#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from datetime import time
from functools import partial

import pytest
from openpyxl import Workbook

from meeting_attendance_workbook import (
    AttendanceInfo, GET_NICKNAME_ERROR, OVERVIEW_OF_MEMBER_ATTENDANCE,
    create_origin_attendance_info, get_nickname, merge_attendance_infos,
    parse_attendance_info, parse_attendance_sheet
)
from meeting_comm import Cell, StatError


def test_get_nickname_01():
    result = get_nickname('noway(noway)')
    assert 'noway' == result


def test_get_nickname_02():
    result = get_nickname('(noway)')
    assert 'noway' == result


def test_get_nickname_03():
    with pytest.raises(StatError) as ex:
        get_nickname('')
    assert GET_NICKNAME_ERROR == str(ex.value)


def create_test_attendance_workbook_01() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = OVERVIEW_OF_MEMBER_ATTENDANCE
    ws.cell(row=9, column=1, value='用户昵称（入会昵称）')
    ws.cell(row=9, column=2, value='首次入会时间')
    ws.cell(row=9, column=3, value='最后退会时间')
    ws.cell(row=9, column=4, value='入会次数')
    ws.cell(row=9, column=5, value='累计参会时长')
    ws.cell(row=9, column=6, value='身份')

    ws.cell(row=10, column=1, value='人员1(中乾0人员1)')
    ws.cell(row=10, column=2, value='2024-01-01 00:00:00')
    ws.cell(row=10, column=3, value='2024-01-01 00:30:00')
    ws.cell(row=10, column=4, value='1')
    ws.cell(row=10, column=5, value='0:30:00')
    ws.cell(row=10, column=6, value='普通参会者')

    ws.cell(row=11, column=1, value='人员2(人员2)')
    ws.cell(row=11, column=2, value='2024-01-01 00:00:00')
    ws.cell(row=11, column=3, value='2024-01-01 00:30:00')
    ws.cell(row=11, column=4, value='2')
    ws.cell(row=11, column=5, value='0:30:00')
    ws.cell(row=11, column=6, value='普通参会者')

    ws.cell(row=12, column=1, value='人员2(人员2)')
    ws.cell(row=12, column=2, value='2024-01-01 00:30:00')
    ws.cell(row=12, column=3, value='2024-01-01 00:40:00')
    ws.cell(row=12, column=4, value='1')
    ws.cell(row=12, column=5, value='0:10:00')
    ws.cell(row=12, column=6, value='普通参会者')

    ws.cell(row=13, column=1, value='人员3&人员4(中乾2人员3＆中坤0人员4)')
    ws.cell(row=13, column=2, value='2024-01-01 00:00:00')
    ws.cell(row=13, column=3, value='2024-01-01 00:40:00')
    ws.cell(row=13, column=4, value='1')
    ws.cell(row=13, column=5, value='0:40:00')
    ws.cell(row=13, column=6, value='主持人')

    return wb


TEST_ATTENDANCE_WB_01 = create_test_attendance_workbook_01()

TEST_ATTENDANCE_INFO_01 = create_origin_attendance_info(
    '中乾0人员1', '人员1(中乾0人员1)', time(0, 30, 0)
)
TEST_ATTENDANCE_INFO_01_01 = create_origin_attendance_info(
    '*中乾0人员1*', '*中乾0人员1*', time(0, 30, 0)
)
TEST_ATTENDANCE_INFO_02 = AttendanceInfo(
    '人员2', '人员2(人员2)', time(0, 40, 0), True
)
TEST_ATTENDANCE_INFO_03 = create_origin_attendance_info(
    '中乾2人员3', '人员3&人员4(中乾2人员3＆中坤0人员4)', time(0, 40, 0)
)
TEST_ATTENDANCE_INFO_04 = create_origin_attendance_info(
    '中坤0人员4', '人员3&人员4(中乾2人员3＆中坤0人员4)', time(0, 40, 0)
)

get_test_attendance_infos_01 = partial(
    parse_attendance_sheet, TEST_ATTENDANCE_WB_01[OVERVIEW_OF_MEMBER_ATTENDANCE]
)


def test_create_origin_attendance_info_01():
    result = create_origin_attendance_info(
        '中乾0人员1', '中乾0人员1', time(0, 30, 0)
    )
    expected = AttendanceInfo('中乾0人员1', '中乾0人员1', time(0, 30, 0), False)
    assert expected == result


def test_parse_attendance_info_01():
    result = list(
        parse_attendance_info(
            Cell('(人员1)'),
            Cell(None),
            Cell(None),
            Cell(None),
            Cell('0:30:00'),
            Cell(None)
        )
    )
    expected = [create_origin_attendance_info('人员1', '(人员1)', time(0, 30, 0))]
    assert expected == result


def test_parse_attendance_info_02():
    result = list(
        parse_attendance_info(
            Cell('(人员 _-01＆人员2)'),
            Cell(None),
            Cell(None),
            Cell(None),
            Cell('0:30:00'),
            Cell(None)
        )
    )
    expected = [
        create_origin_attendance_info('人员1', '(人员 _-01＆人员2)', time(0, 30, 0)),
        create_origin_attendance_info('人员2', '(人员 _-01＆人员2)', time(0, 30, 0)),
    ]
    assert expected == result


def test_parse_attendance_info_03():
    result = list(
        parse_attendance_info(
            Cell('(人员1&人员2)'),
            Cell(None),
            Cell(None),
            Cell(None),
            Cell('0:30:00'),
            Cell(None)
        )
    )
    expected = [
        create_origin_attendance_info('人员1', '(人员1&人员2)', time(0, 30, 0)),
        create_origin_attendance_info('人员2', '(人员1&人员2)', time(0, 30, 0)),
    ]
    assert expected == result


def test_merge_attendance_infos_01():
    result = merge_attendance_infos((
        create_origin_attendance_info(
            '中乾0人员1', '中乾0人员1', time(0, 30, 30)
        ),
        create_origin_attendance_info(
            '中乾0人员1', '中乾0人员1', time(0, 30, 41)
        ),
    ))
    expected = AttendanceInfo(
        '中乾0人员1', '中乾0人员1', time(1, 1, 11), True
    )
    assert expected == result


def test_parse_attendance_sheet_01():
    result = parse_attendance_sheet(
        TEST_ATTENDANCE_WB_01[OVERVIEW_OF_MEMBER_ATTENDANCE]
    )
    expected = (
        TEST_ATTENDANCE_INFO_01,
        TEST_ATTENDANCE_INFO_03,
        TEST_ATTENDANCE_INFO_04,
        TEST_ATTENDANCE_INFO_02,
    )
    assert expected == result
