#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""生活修行考勤表。"""

import os
import re
from argparse import Namespace
from datetime import datetime, time, timedelta
from functools import partial
from itertools import (
    chain, filterfalse, groupby, repeat
)
from operator import (
    attrgetter, eq, itemgetter, lt, methodcaller
)
from typing import Dict, Iterator, List, NamedTuple, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, Side
from pypinyin import pinyin, Style

from meeting_attendance_workbook import (
    AttendanceInfo, AttendanceInfos, DETAIL_OF_MEMBER_ATTENDANCE,
    does_attendance_detail_info_intersect,
    normalize_attendance_detail_info_time, merge_attendance_infos,
    parse_attendance_detail_sheet, partition_attendance_infos,
    summarize_attendance_time,
)
from meeting_comm import (
    MEETING_SUMMARY_FILENAME, MEETING_SUMMARY_OUTPUT_FILENAME,
    MEETING_ATTENDANCE_FILENAME,
    constant, cross, dispatch, ensure, identity, if_, invoke, pipe,
    side_effect, starapply, to_stream, tuple_args,
    dict_groupby, expand_groupby,
)


PEOPLE_SHEET_NAME = '人员总表'
MEETING_INFO_SHEET_NAME = '参数'
MISMATCHED_SHEET_NAME = '未改名'
TOTAL_ABSENT_SHEET_NAME = '缺勤总表'
TEAM_MAPPING_SHEET_NAME = '小组映射表'

TOTAL_ABSENT_SHEET_FIRST_LINE = 3

TEAM_NAME_REGEX = re.compile(r'(..)组')

BLACK_FONT = Font(color='00000000')
RED_FONT = Font(color='00FF0000')
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
BORDER_SIDE = Side(border_style='thin', color='00000000')
BORDER = Border(
    left=BORDER_SIDE, right=BORDER_SIDE,
    top=BORDER_SIDE, bottom=BORDER_SIDE,
)


class PersoneelInfo(NamedTuple):
    """人员信息。"""
    name: str  # 姓名
    team: str  # 小组
    number: int  # 小组编号
    
    @property
    def team_number(self) -> str:
        """小组名+编号。"""
        return f'{self.team}{self.number}'

    @property
    def formal_name(self) -> str:
        """正式名称。"""
        return f'{self.team}{self.number}{self.name}'

    @property
    def formal_nick_name(self) -> str:
        """正式昵称。"""
        if len(self.name) >= 3:
            nick_name = self.name[1:]
        else:
            nick_name = self.name
        return f'{self.team}{self.number}{nick_name}'

    @property
    def formal_pinyin_name(self) -> str:
        """正式拼音名称。"""
        pinyin_team = ''.join(chain.from_iterable(pinyin(self.team, style=Style.FIRST_LETTER)))
        pinyin_team = pinyin_team.upper()
        return f'{pinyin_team}{self.number}{self.name}'


PersoneelInfos = Tuple[PersoneelInfo, ...]


class PersoneelAttendanceInfo(NamedTuple):
    """个人参会信息。"""
    personeel_info: PersoneelInfo
    personeel_attendance_infos: AttendanceInfos
    personeel_attendance_time: timedelta
    is_attendanced: bool


PersoneelAttendanceInfos = Tuple[PersoneelAttendanceInfo, ...]

TeamAttendanceInfos = dict[str, Tuple[PersoneelAttendanceInfo, ...]]

ZoneAttendanceInfos = dict[str, TeamAttendanceInfos]


class MeetingInfo(NamedTuple):
    """会议信息。"""
    solar_term: str
    meeting_start_time: datetime
    meeting_end_time: datetime
    meeting_time: int
    meeting_enough_time: int


class FillCommand(NamedTuple):
    """填充指令。"""
    line_no: int
    column_no: int
    text: str
    is_red: bool  # 是否为红字
    is_absent: bool = False


MEETING_INFO_KEYS = {
    '节气名': 'solar_term',
    '会议开始时间': 'meeting_start_time',
    '会议结束时间': 'meeting_end_time',
    '会议总时长': 'meeting_time',
}


def parse_personnel_info(row: Tuple[str, ...]) -> PersoneelInfo:
    """解析人员信息。"""
    return PersoneelInfo(row[0], row[2], int(row[3]))


# 转换人员总表为内部数据结构
# Worksheet -> Tuple[Tuple[str, ...], ...]
convert_people_sheet = pipe(
    methodcaller('iter_rows', min_row=2, min_col=2, max_col=5, values_only=True),
    tuple,
)

# 解析人员总表
# Worksheet -> Tuple[PersoneelInfos, Tuple[int, str]]
parse_people_sheet = pipe(
    convert_people_sheet,
    partial(filter, pipe(itemgetter(0), bool)),
    partial(map, parse_personnel_info),
    tuple
)

# 转换会议信息为内部数据结构
# Worksheet -> Tuple[Tuple[Union[str, datetime], ...], ...]
convert_meeting_info_sheet = pipe(
    methodcaller('iter_rows', min_row=1, max_col=2, values_only=True),
    tuple,
)

# 转换小组映射为内部数据结构
# Worksheet -> Tuple[Tuple[str, ...], ...]
convert_team_mapping_sheet = pipe(
    methodcaller('iter_rows', min_row=1, max_col=2, values_only=True),
    tuple,
)


# 解析会议信息
# Tuple[Any, ...] -> Tuple[str, Any]
parse_meeting_info = pipe(
    cross(
        pipe(
            MEETING_INFO_KEYS.get,
            partial(ensure, bool),
        ),
        identity,
    ),
    tuple,
)


def parse_meeting_info_sheet(sheet) -> MeetingInfo:
    """解析会议信息工作表。"""
    info_dict = dict(map(parse_meeting_info, convert_meeting_info_sheet(sheet)))
    meeting_time: timedelta = info_dict['meeting_end_time'] - info_dict['meeting_start_time']
    info_dict['meeting_time'] = int(meeting_time.total_seconds()) // 60
    info_dict['meeting_enough_time'] = info_dict['meeting_time'] * 2 // 3
    return MeetingInfo(**info_dict)


def parse_team_mapping_sheet(sheet) -> Dict[str, str]:
    """解析小组映射工作表。"""
    team_mapping = dict(convert_team_mapping_sheet(sheet))
    return team_mapping


def timedelta_to_time(td: timedelta) -> time:
    """时间差转为时间。"""
    return time(
        td.seconds // 3600, (td.seconds // 60) % 60, td.seconds % 60
    )


def generate_mismatched_command(idx: int,
                                pair: Tuple[str, AttendanceInfos]
                                ) -> Iterator[FillCommand]:
    """生成没有匹配的出席信息表填充指令。"""
    origin_name, attendance_time = pair
    yield FillCommand(idx, 1, origin_name, False)
    yield FillCommand(idx, 2, attendance_time.strftime('%H:%M:%S'), False)


# 生成没有匹配的出席信息表填充指令
# Iterator[Tuple[str, AttendanceInfos], ...] -> Tuple[FillCommand, ...]
generate_mismatched_commands = pipe(
    partial(map,
        pipe(
            if_(
                pipe(itemgetter(0), partial(eq, '(None)')),
                pipe(
                    cross(
                        repeat,
                        partial(
                            map,
                            pipe(
                                to_stream,
                                summarize_attendance_time,
                                timedelta_to_time
                            )
                        ),
                    ),
                    starapply(zip),
                    tuple,
                ),
                pipe(
                    cross(
                        identity,
                        pipe(summarize_attendance_time, timedelta_to_time),
                    ),
                    to_stream,
                ),
            ),
        ),
    ),
    chain.from_iterable,
    partial(enumerate, start=2),
    partial(map, starapply(generate_mismatched_command)),
    chain.from_iterable,
    tuple,
)

# 填充单元格
# Tuple[WorksheetCell, FillCommand] -> None
fill_workcell = pipe(
    tuple_args,
    side_effect(
        pipe(
            dispatch(
                itemgetter(0),
                constant('value'),
                pipe(itemgetter(1), itemgetter(2)),
            ),
            starapply(setattr),
        ),
    ),
    side_effect(
        if_(
            pipe(itemgetter(1), itemgetter(3)),
            pipe(
                dispatch(
                    itemgetter(0),
                    constant('font'),
                    constant(RED_FONT),
                ),
                starapply(setattr),
            ),
            pipe(
                dispatch(
                    itemgetter(0),
                    constant('font'),
                    constant(BLACK_FONT),
                ),
                starapply(setattr),
            )
        ),
    ),
    side_effect(
        if_(
            pipe(
                dispatch(
                    pipe(itemgetter(1), len, partial(lt, 4)),
                    pipe(itemgetter(1), itemgetter(4)),
                ),
                all,
            ),
            pipe(
                side_effect(
                    pipe(
                        dispatch(
                            itemgetter(0),
                            constant('alignment'),
                            constant(CENTER_ALIGN),
                        ),
                        starapply(setattr),
                    )
                ),
                side_effect(
                    pipe(
                        dispatch(
                            itemgetter(0),
                            constant('border'),
                            constant(BORDER),
                        ),
                        starapply(setattr),
                    )
                ),
            ),
        ),
    ),
)

# 填充工作表
# Tuple[Worksheet, FillCommand] -> None
fill_worksheet_command = pipe(
    tuple_args,
    dispatch(
        pipe(
            dispatch(
                pipe(itemgetter(0), attrgetter('cell')),
                pipe(itemgetter(1), itemgetter(0)),
                pipe(itemgetter(1), itemgetter(1)),
            ),
            starapply(invoke),
        ),
        itemgetter(1),
    ),
    tuple,
    fill_workcell,
)

# 填充工作表
# Tuple[Worksheet, Tuple[FillCommand, ...]] -> Tuple[FillCommand, ...]
do_fill_worksheet_commands = pipe(
    tuple_args,
    side_effect(
        pipe(
            cross(repeat, identity),
            starapply(zip),
            partial(map, fill_worksheet_command),
            tuple,
        ),
    ),
    itemgetter(1),
)


def normalize_attendance_detail_infos(meeting_info: MeetingInfo):
    """标准化参会明细信息。"""
    return pipe(
        partial(
            filter,
            partial(
                does_attendance_detail_info_intersect,
                meeting_info.meeting_start_time,
                meeting_info.meeting_end_time,
            ),
        ),
        partial(
            map,
            partial(
                normalize_attendance_detail_info_time,
                meeting_info.meeting_start_time,
                meeting_info.meeting_end_time
            )
        ),
        tuple,
    )

class StatAttendanceInfos:

    def __init__(self, name_match: bool = False):
        self.name_match = name_match

    def match_personeel_info_and_attendance_info(self,
                                                 personeel_info: PersoneelInfo,
                                                 attendance_info: AttendanceInfo) -> bool:
        """匹配个人信息和参会信息。"""
        if personeel_info.formal_name in attendance_info.nickname:
            return True
        if personeel_info.formal_name in attendance_info.meeting_name:
            return True
        if personeel_info.formal_nick_name in attendance_info.nickname:
            return True
        if personeel_info.formal_nick_name in attendance_info.meeting_name:
            return True
        if personeel_info.formal_pinyin_name in attendance_info.nickname:
            return True
        if personeel_info.formal_pinyin_name in attendance_info.meeting_name:
            return True

        if personeel_info.team_number == attendance_info.nickname:
            return True

        if self.name_match:
            if personeel_info.name == attendance_info.nickname:
                return True
            for idx in range(1, min(len(attendance_info.nickname)-1, 3)):
                if personeel_info.name == attendance_info.nickname[idx:]:
                    return True

        return False


    def stat_personeel_attendance_infos(self,
                                        personeel_info: PersoneelInfo,
                                        attendance_infos: dict[str, AttendanceInfos],
                                        ) -> Iterator[AttendanceInfo]:
        """统计个人参会详情。"""
        for _, one_attendance_infos in attendance_infos.items():
            matched = any(
                map(
                    partial(self.match_personeel_info_and_attendance_info, personeel_info),
                    one_attendance_infos
                )
            )
            if matched:
                yield from one_attendance_infos


    def stat_people_attendance_infos(self,
                                     personeel_infos: PersoneelInfos,
                                     attendance_infos: dict[str, AttendanceInfos],
                                     meeting_info: MeetingInfo,
                                     ) -> Iterator[PersoneelAttendanceInfo]:
        """统计个人参会详情。"""
        enough_attendance_time = timedelta(minutes=meeting_info.meeting_enough_time)

        for personeel_info in personeel_infos:
            personeel_attendance_infos = tuple(
                self.stat_personeel_attendance_infos(personeel_info, attendance_infos)
            )
            personeel_attendance_time = summarize_attendance_time(
                normalize_attendance_detail_infos(meeting_info)(personeel_attendance_infos)
            )
            is_attendanced = personeel_attendance_time >= enough_attendance_time
            yield PersoneelAttendanceInfo(
                personeel_info, personeel_attendance_infos, personeel_attendance_time,
                is_attendanced
            )


def stat_mismatched_attendance_infos(matched_attendance_infos: AttendanceInfos,
                                     attendance_infos: dict[str, AttendanceInfos]
                                     ) -> Iterator[AttendanceInfo]:
    """统计没有匹配的参会信息。"""
    for attendance_info in chain.from_iterable(attendance_infos.values()):
        if attendance_info not in matched_attendance_infos:
            yield attendance_info


def classify_team_attendance_infos(people_attendance_infos: PersoneelAttendanceInfos,
                                   ) -> TeamAttendanceInfos:
    """分类小组参会信息。"""
    return dict(
        expand_groupby(
            groupby(
                people_attendance_infos,
                key=pipe(itemgetter(0), attrgetter('team'))
            )
        )
    )


def classify_zone_attendance_infos(team_attendance_infos: TeamAttendanceInfos,
                                   team_mapping: dict[str, str]
                                   ) -> ZoneAttendanceInfos:
    """分类区域参会信息。"""
    return dict(
        dict_groupby(
            groupby(
                team_attendance_infos.items(),
                key=pipe(itemgetter(0), team_mapping.get)
            )
        )
    )


def generate_attendance_infos_fill_commands(team_attendance_infos: TeamAttendanceInfos
                                            ) -> Iterator[FillCommand]:
    """生成参会信息的填充命令。"""
    for idx, items in enumerate(team_attendance_infos.items()):
        team, personeel_attendance_infos = items
        start_line = idx * 11 + 1
        yield FillCommand(
            start_line, 1, f'{team}组（{len(personeel_attendance_infos)}人）', False
        )
        yield FillCommand(start_line, 2, '序号', False)
        yield FillCommand(start_line, 3, '用户入会昵称', False)
        yield FillCommand(start_line, 4, '累计参会时长', False)

        absent_attendance_infos = tuple(
            filterfalse(
                attrgetter('is_attendanced'), personeel_attendance_infos
            )
        )

        if not absent_attendance_infos:
            yield FillCommand(start_line + 1, 1, '全勤', False)

        for absent_idx, absent_attendance_info in enumerate(absent_attendance_infos, start=1):
            yield FillCommand(
                start_line + absent_idx, 1,
                absent_attendance_info.personeel_info.formal_name, False
            )

        attendanced_attendance_infos = tuple(
            filter(
                attrgetter('personeel_attendance_infos'), personeel_attendance_infos
            )
        )
        for attendanced_idx, attendanced_attendance_info in enumerate(attendanced_attendance_infos, start=1):
            yield FillCommand(
                start_line + attendanced_idx, 2,
                attendanced_idx, False
            )
            yield FillCommand(
                start_line + attendanced_idx, 3,
                attendanced_attendance_info.personeel_info.formal_name, False
            )
            attendance_time = timedelta_to_time(
                attendanced_attendance_info.personeel_attendance_time
            ).strftime('%H:%M:%S')
            yield FillCommand(
                start_line + attendanced_idx, 4,
                attendance_time,
                not attendanced_attendance_info.is_attendanced
            )
        not_attendanced_attendance_infos = tuple(
            filterfalse(
                attrgetter('personeel_attendance_infos'), personeel_attendance_infos
            )
        )
        for attendanced_idx, attendanced_attendance_info in enumerate(not_attendanced_attendance_infos, start=len(attendanced_attendance_infos)+1):
            yield FillCommand(
                start_line + attendanced_idx, 2,
                attendanced_idx, False
            )
            yield FillCommand(
                start_line + attendanced_idx, 3,
                attendanced_attendance_info.personeel_info.formal_name, False
            )
            yield FillCommand(
                start_line + attendanced_idx, 4, '缺席', True,
            )


def fill_mismatched_attendance_infos(mismatched_attendance_infos: AttendanceInfos,
                                     summary_workbook: Workbook) -> Tuple[FillCommand, ...]:
    """填充未改名参会信息。"""
    mismatched_attendance_infos = merge_attendance_infos(
      mismatched_attendance_infos
    ).items()
    mismatched_commands = generate_mismatched_commands(mismatched_attendance_infos)
    mismatched_sheet = summary_workbook[MISMATCHED_SHEET_NAME]
    fill_mismatched_commands = do_fill_worksheet_commands(mismatched_sheet, mismatched_commands)
    return fill_mismatched_commands


def fill_zone_attendance_infos(zone_attendance_infos: ZoneAttendanceInfos,
                               workbook: Workbook):
    """填充区域的参会信息。"""
    for zone, attendance_infos in zone_attendance_infos.items():
        fill_commands = generate_attendance_infos_fill_commands(attendance_infos)
        do_fill_worksheet_commands(workbook[zone], fill_commands)


def overlapped(items: List) -> bool:
    """是否重叠。"""
    result = False
    for idx_i in range(len(items)):
        for idx_j in range(idx_i + 1, len(items)):
            if items[idx_i] == items[idx_j]:
                result = True
    return result


def stat_time(args: Namespace) -> bool:
    """统计参会时长。"""
    summary_workbook = load_workbook(
        os.path.join(args.meeting, MEETING_SUMMARY_FILENAME)
    )
    attendance_workbook = load_workbook(
        os.path.join(args.meeting, MEETING_ATTENDANCE_FILENAME)
    )
    summary_workbook_output_filepath = os.path.join(
        args.meeting, MEETING_SUMMARY_OUTPUT_FILENAME
    )
    people_sheet = summary_workbook[PEOPLE_SHEET_NAME]
    personeel_infos = parse_people_sheet(people_sheet)
    attendance_infos = parse_attendance_detail_sheet(
        attendance_workbook[DETAIL_OF_MEMBER_ATTENDANCE]
    )
    meeting_info = parse_meeting_info_sheet(summary_workbook[MEETING_INFO_SHEET_NAME])
    team_mapping = parse_team_mapping_sheet(summary_workbook[TEAM_MAPPING_SHEET_NAME])

    print(f'会议时长为{meeting_info.meeting_time}分钟。')
    print(f'参会时间下限为{meeting_info.meeting_enough_time}分钟。')

    attendance_infos = partition_attendance_infos(attendance_infos)

    people_attendance_infos = tuple(
        StatAttendanceInfos(
            not overlapped(list(map(attrgetter('name'), personeel_infos)))
        ).stat_people_attendance_infos(
            personeel_infos, attendance_infos, meeting_info,
        )
    )

    matched_attendance_infos = set(
        chain.from_iterable(
            map(attrgetter('personeel_attendance_infos'), people_attendance_infos)
        )
    )

    team_attendance_infos = classify_team_attendance_infos(people_attendance_infos)
    zone_attendance_infos = classify_zone_attendance_infos(
        team_attendance_infos, team_mapping
    )

    mismatched_attendance_infos = tuple(
        sorted(
            normalize_attendance_detail_infos(meeting_info)(
                stat_mismatched_attendance_infos(
                    matched_attendance_infos, attendance_infos
                )
            ),
            key=itemgetter(0)
        )
    )

    fill_mismatched_attendance_infos(mismatched_attendance_infos, summary_workbook)
    fill_zone_attendance_infos(zone_attendance_infos, summary_workbook)

    summary_workbook.save(summary_workbook_output_filepath)
    print(f"保存'{summary_workbook_output_filepath}'文件成功。")

    return True


def main_process(args: Namespace):
    """主流程。"""
    stat_time(args)
